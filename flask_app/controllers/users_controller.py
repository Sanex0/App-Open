import smtplib
import json
import requests
import pprint
import random
from decimal import Decimal
from email.message import EmailMessage
from datetime import datetime

from flask import render_template, redirect, request, session, flash, url_for, jsonify, Response
from flask_bcrypt import Bcrypt

from flask_app import app
from flask_app.models.productos import Producto
from flask_app.models.users import User
from flask_app.models.cajas import Caja
from flask_app.models.apertura import Apertura
from flask_app.models.venta import Venta
from flask_app.models.permiso import Permiso
from flask_app.config.conexiones import connectToMySQL

bcrypt = Bcrypt(app)

# --- Funciones Auxiliares ---

def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, Producto):
        return obj.__dict__
    raise TypeError

def validar_rut(rut):
    if not rut: return False
    rut = rut.replace('.', '').replace('-', '').upper()
    if len(rut) < 8 or len(rut) > 9:
        return False
    cuerpo = rut[:-1]
    dv = rut[-1]
    if not cuerpo.isdigit() or not (dv.isdigit() or dv == 'K'):
        return False
    
    suma = sum(int(c) * (multiplo if multiplo < 8 else multiplo - 6) for c, multiplo in zip(reversed(cuerpo), range(2, len(cuerpo) + 2)))
    dv_esperado = 11 - (suma % 11)
    if dv_esperado == 11: dv_esperado = '0'
    elif dv_esperado == 10: dv_esperado = 'K'
    else: dv_esperado = str(dv_esperado)
    
    return dv == dv_esperado


def send_email(to_address, subject, body, html_body=None, sender=None, attachments=None):
    """Enviar un email simple usando la configuración en app.config.
    Lanza excepciones en caso de error para que el llamador las maneje.
    attachments: lista de dicts con keys 'filename', 'content' (bytes), 'maintype', 'subtype'
    """
    # Credenciales SMTP múltiples para balanceo de carga (cPanel)
    # Cada correo usa su contraseña correspondiente según el índice
    smtp_users = [
        "no-responder@clubrecrear.cl",
        "no-responder1@clubrecrear.cl",
        "no-responder2@clubrecrear.cl"
    ]
    smtp_passwords = [
        "++&x,TyMji!;",
        "L7MVjISZvw1t",
        "3RJC^Of(L_t}"
    ]
    
    # Seleccionar credenciales aleatorias (correo y su contraseña correspondiente por índice)
    indice = random.randint(0, len(smtp_users) - 1)
    mail_user = smtp_users[indice]
    mail_pass = smtp_passwords[indice]
    
    mail_server = app.config.get('MAIL_SERVER', 'mail.clubrecrear.cl')
    mail_port = int(app.config.get('MAIL_PORT', 465) or 465)
    mail_use_tls = app.config.get('MAIL_USE_TLS', True)
    mail_use_ssl = app.config.get('MAIL_USE_SSL', False)
    sender = sender or app.config.get('MAIL_DEFAULT_SENDER', None)
    # If we have SMTP credentials, prefer using the authenticated user as sender
    effective_sender = mail_user or sender or f'no-reply@{mail_server}'

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = effective_sender
    msg['To'] = to_address
    msg.set_content(body)
    
    if html_body:
        msg.add_alternative(html_body, subtype='html')

    if attachments:
        for att in attachments:
            msg.add_attachment(
                att['content'],
                maintype=att.get('maintype', 'application'),
                subtype=att.get('subtype', 'octet-stream'),
                filename=att.get('filename', 'attachment')
            )

    if app.config.get('LOGIN_DEBUG'):
        try:
            print(f"[EMAIL DEBUG] server={mail_server} port={mail_port} use_tls={mail_use_tls} use_ssl={mail_use_ssl} user_set={'yes' if mail_user else 'no'} sender={effective_sender}")
        except Exception:
            pass

    if mail_use_ssl:
        with smtplib.SMTP_SSL(mail_server, mail_port) as server:
            if mail_user and mail_pass:
                server.login(mail_user, mail_pass)
            server.send_message(msg)
    else:
        with smtplib.SMTP(mail_server, mail_port) as server:
            if mail_use_tls:
                server.starttls()
            if mail_user and mail_pass:
                server.login(mail_user, mail_pass)
            server.send_message(msg)

# --- Rutas ---

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    if not request.form.get('email') or not request.form.get('password'):
        flash('Faltan campos obligatorios', 'inicio_sesion')
        return redirect('/')
    email = request.form['email']
    pwd = request.form['password']

    user = User.get_by_email({'email_usuario': email})

    # Logs de diagnóstico controlados por la configuración `LOGIN_DEBUG`.
    if app.config.get('LOGIN_DEBUG'):
        if not user:
            print(f"[LOGIN DEBUG] No se encontró usuario para email: {email}")
        else:
            # stored = user.password or ''
            # prefix = stored[:6] if len(stored) >= 6 else stored
            # print(f"[LOGIN DEBUG] Usuario encontrado id={getattr(user, 'id_usuario', None)}; hash_len={len(stored)}; hash_prefix={prefix}")
            pass

    # Diferenciar mensajes: usuario inexistente vs contraseña incorrecta.
    # Para evitar *user enumeration* en producción, mostramos un mensaje genérico
    # a menos que `LOGIN_DEBUG` esté activado.
    if not user:
        if app.config.get('LOGIN_DEBUG'):
            flash('No existe un usuario con ese correo', 'inicio_sesion')
        else:
            flash('Credenciales incorrectas', 'inicio_sesion')
        return redirect('/')

    # Verificar contraseña con manejo seguro de errores en caso de hash inválido
    try:
        valid_password = bcrypt.check_password_hash(user.password, pwd)
    except ValueError as e:
        # Ocurre cuando el valor almacenado no es un salt bcrypt válido
        if app.config.get('LOGIN_DEBUG'):
            stored = getattr(user, 'password', None)
            print(f"[LOGIN DEBUG] bcrypt error for user id={getattr(user,'id_usuario', None)}: {e}; stored_hash={repr(stored)[:200]}")
            flash('Error en el formato del hash de la contraseña. Contacte al administrador.', 'inicio_sesion')
        else:
            flash('Credenciales incorrectas', 'inicio_sesion')
        return redirect('/')

    if not valid_password:
        if app.config.get('LOGIN_DEBUG'):
            flash('Contraseña incorrecta', 'inicio_sesion')
        else:
            flash('Credenciales incorrectas', 'inicio_sesion')
        return redirect('/')
        
    session['user_id'] = user.id_usuario
    return redirect('/index.html')

@app.route('/index.html')
def index_html():
    if 'user_id' not in session:
        return redirect('/')
    
    user = User.get_by_id(session['user_id'])
    
    permisos = Permiso.get_by_user_id(session['user_id'])
    allowed_caja_ids = [p.vta_cajas_id_caja for p in permisos]
    
    cajas = Caja.get_all()
    allowed_cajas = [c for c in cajas if c.id_caja in allowed_caja_ids]
    # if app.config.get('LOGIN_DEBUG'):
    #     try:
    #         print(f"[LOGIN DEBUG] user.id_usuario={getattr(user,'id_usuario', None)}")
    #         print(f"[LOGIN DEBUG] permisos raw count={len(permisos)}")
    #         for i, p in enumerate(permisos):
    #             print(f"[LOGIN DEBUG] permiso[{i}] = {{'id_permiso': getattr(p,'id_permiso', None), 'id_usuario_fk': getattr(p,'id_usuario_fk', None), 'vta_cajas_id_caja': getattr(p,'vta_cajas_id_caja', None)}}")
    #         print(f"[LOGIN DEBUG] allowed_caja_ids = {allowed_caja_ids}")
    #         print(f"[LOGIN DEBUG] total cajas from DB = {len(cajas)}")
    #         for c in cajas[:20]:
    #             print(f"[LOGIN DEBUG] caja: id_caja={getattr(c,'id_caja', None)}, detalle_caja={getattr(c,'detalle_caja', None)}")
    #         print(f"[LOGIN DEBUG] allowed_cajas_count = {len(allowed_cajas)}")
    #     except Exception as e:
    #         print(f"[LOGIN DEBUG] error printing debug info: {e}")
    
    return render_template('index.html', user=user, cajas=allowed_cajas)

@app.route('/caja/<int:id_caja>')
def ver_caja(id_caja):
    if 'user_id' not in session:
        return redirect('/')
    
    # Obtener información de la caja actual
    current_caja = Caja.get_by_id(id_caja)
    is_variable = current_caja.es_variable if current_caja else 0
    
    productos = Producto.get_by_caja(id_caja)
    cajas = Caja.get_all()
    # Apertura activa (por caja)
    active_apertura = Apertura.get_active_by_caja(id_caja)
    apertura_totals = 0
    if active_apertura:
        try:
            apertura_totals = Apertura.get_totals_for_apertura(active_apertura.id_apertura)
        except Exception as e:
            apertura_totals = 0
    # Permisos del usuario sobre cajas (para mostrar mensajes útiles)
    try:
        permisos = Permiso.get_by_user_id(session['user_id'])
        allowed_caja_ids = [p.vta_cajas_id_caja for p in permisos]
    except Exception:
        allowed_caja_ids = []
    can_open_apertura = (id_caja in allowed_caja_ids)
    # Restaurar cantidades previamente seleccionadas (si existen) desde la sesión
    try:
        session_products = session.get('productos_boleta', []) or []
        # session_products es una lista de dicts con 'id_producto' y 'cantidad'
        qty_map = {str(p.get('id_producto')): int(p.get('cantidad', 0)) for p in session_products}
        for prod in productos:
            pid = str(getattr(prod, 'id_producto', ''))
            prod.cantidad = qty_map.get(pid, 0)
    except Exception as e:
        if app.config.get('LOGIN_DEBUG'):
            print(f"[LOGIN DEBUG] Error al restaurar cantidades desde session: {e}")

    return render_template('caja.html', productos=productos, id_caja=id_caja, cajas=cajas, apertura=active_apertura, apertura_totals=apertura_totals, can_open_apertura=can_open_apertura, current_caja=current_caja, is_variable=is_variable)


@app.route('/aperturas')
def listar_aperturas():
    if 'user_id' not in session:
        return redirect('/')

    permisos = Permiso.get_by_user_id(session['user_id'])
    allowed_caja_ids = [p.vta_cajas_id_caja for p in permisos]
    # obtener aperturas para las cajas permitidas
    try:
        aperturas = Apertura.get_all_by_cajas(allowed_caja_ids)
    except Exception as e:
        if app.config.get('LOGIN_DEBUG'):
            print(f"[APERTURAS DEBUG] Error obteniendo aperturas: {e}")
        aperturas = []

    # obtener información de cajas para mostrar el nombre
    all_cajas = Caja.get_all()
    cajas = {c.id_caja: c for c in all_cajas}
    # cajas permitidas para este usuario
    allowed_cajas = [c for c in all_cajas if c.id_caja in allowed_caja_ids]

    # calcular totales de ventas por apertura para mostrar en la tabla
    totals_map = {}
    try:
        for ap in aperturas:
            try:
                totals_map[ap.id_apertura] = Apertura.get_totals_for_apertura(ap.id_apertura)
            except Exception:
                totals_map[ap.id_apertura] = 0
    except Exception:
        totals_map = {}

    # --- KPIs para los chips (ventas del día, total del día, arqueos abiertos, último arqueo) ---
    ventas_count = 0
    ventas_total_hoy = 0
    arqueos_abiertos_count = 0
    ultimo_arqueo_creado = None

    try:
        if allowed_caja_ids:
            placeholders = ','.join(['%s'] * len(allowed_caja_ids))
            # Contar ventas del día (usamos la fecha de la apertura asociada como aproximación)
            q_ventas = (
                f"SELECT COUNT(v.id_ventas) AS cnt, IFNULL(SUM(v.total_ventas),0) AS total "
                f"FROM vta_ventas v JOIN vta_apertura a ON v.id_apertura = a.id_apertura "
                f"WHERE a.id_caja_fk IN ({placeholders}) AND DATE(a.fecha_inicio_apertura) = CURDATE();"
            )
            res_v = connectToMySQL('sistemas').query_db(q_ventas, tuple(allowed_caja_ids))
            if res_v and isinstance(res_v, list) and len(res_v) > 0:
                ventas_count = int(res_v[0].get('cnt', 0) or 0)
                ventas_total_hoy = int(res_v[0].get('total', 0) or 0)

            # contar aperturas abiertas para esas cajas
            q_abiertos = f"SELECT COUNT(*) AS cnt FROM vta_apertura WHERE estado_apertura = 1 AND id_caja_fk IN ({placeholders});"
            res_a = connectToMySQL('sistemas').query_db(q_abiertos, tuple(allowed_caja_ids))
            if res_a and isinstance(res_a, list) and len(res_a) > 0:
                arqueos_abiertos_count = int(res_a[0].get('cnt', 0) or 0)

            # obtener la fecha del último arqueo creado (más reciente fecha_inicio_apertura)
            q_ultimo = f"SELECT fecha_inicio_apertura FROM vta_apertura WHERE id_caja_fk IN ({placeholders}) ORDER BY fecha_inicio_apertura DESC LIMIT 1;"
            res_u = connectToMySQL('sistemas').query_db(q_ultimo, tuple(allowed_caja_ids))
            if res_u and isinstance(res_u, list) and len(res_u) > 0:
                ultimo_arqueo_creado = res_u[0].get('fecha_inicio_apertura')
    except Exception as e:
        if app.config.get('LOGIN_DEBUG'):
            print(f"[APERTURAS KPI DEBUG] Error calculando KPIs: {e}")

    most_recent = aperturas[0] if aperturas else None
    # Si venimos de cerrar una apertura, se guarda en session para mostrar el resumen modal
    apertura_resumen_id = None
    try:
        apertura_resumen_id = session.pop('apertura_resumen_id', None)
    except Exception:
        apertura_resumen_id = None

    return render_template('aperturas.html', aperturas=aperturas, cajas=cajas, totals_map=totals_map, allowed_cajas=allowed_cajas, most_recent=most_recent, apertura_resumen_id=apertura_resumen_id, ventas_count=ventas_count, arqueos_abiertos_count=arqueos_abiertos_count, ultimo_arqueo_creado=ultimo_arqueo_creado)


@app.route('/api/caja/<int:id_caja>/productos')
def api_productos_por_caja(id_caja):
    """Endpoint JSON para listar los productos de una caja (útil para pruebas)."""
    if 'user_id' not in session:
        return jsonify({'error': 'not_authenticated'}), 401

    try:
        productos = Producto.get_by_caja(id_caja)
    except Exception as e:
        if app.config.get('LOGIN_DEBUG'):
            print(f"[API DEBUG] Error obteniendo productos para caja {id_caja}: {e}")
        return jsonify({'error': 'internal_error'}), 500

    productos_json = []
    for p in productos:
        productos_json.append({
            'id_prod': getattr(p, 'id_prod', None),
            'id_producto': getattr(p, 'id_producto', None),
            'nombre': getattr(p, 'nombre', None),
            'precio': getattr(p, 'precio', None),
            'compuesto': getattr(p, 'compuesto', None),
            'kitvirtual': getattr(p, 'kitvirtual', None),
        })

    return jsonify(productos_json)

@app.route('/pago', methods=['POST'])
def resumen_pago():
    if 'user_id' not in session:
        return redirect('/')
    
    productos_seleccionados_ids = request.form.getlist('productos')
    if not productos_seleccionados_ids:
        flash('No seleccionaste ningún producto.', 'pago')
        return redirect(request.referrer)
    
    id_caja = request.form.get('id_caja')
    # Verificar que la caja tenga una apertura activa antes de procesar el pago
    try:
        id_caja_int = int(id_caja)
    except Exception:
        flash('Caja inválida.', 'danger')
        return redirect(url_for('listar_aperturas'))

    active_ap = Apertura.get_active_by_caja(id_caja_int)
    if not active_ap:
        flash('No hay una apertura activa para esta caja. Abra una en Gestión de Aperturas antes de procesar pagos.', 'warning')
        return redirect(url_for('listar_aperturas'))

    # Verificar si es caja de precio variable
    precio_variable = request.form.get('precio_variable')
    
    if precio_variable:
        # Manejo para caja de precio variable
        productos_en_caja = Producto.get_by_caja(id_caja_int)
        if not productos_en_caja:
            flash('No hay productos asignados a esta caja.', 'danger')
            return redirect(url_for('ver_caja', id_caja=id_caja_int))
        
        prod = productos_en_caja[0]  # Usar el primer (y único) producto
        
        # El id_producto que viene del form puede ser diferente al que está en prod
        # Buscar la cantidad en el form usando cualquier campo que empiece con 'cantidad_'
        cantidad = 1
        for key in request.form.keys():
            if key.startswith('cantidad_'):
                cantidad = int(request.form.get(key, 1))
                break
        
        precio_unitario = float(precio_variable)
        
        # Calcular total
        total = precio_unitario * cantidad
        
        # Asignar valores al producto
        prod.cantidad = cantidad
        prod.precio = precio_unitario  # Precio unitario para el detalle
        
        productos_a_pagar = [prod]
        
        # Debug
        if app.config.get('LOGIN_DEBUG'):
            print(f"[PRECIO VARIABLE] precio_unitario={precio_unitario}, cantidad={cantidad}, total={total}")
    else:
        # Flujo normal
        productos_en_caja = Producto.get_by_caja(id_caja_int)
        
        productos_a_pagar = []
        total = 0
        
        for prod in productos_en_caja:
            if str(prod.id_producto) in productos_seleccionados_ids:
                cantidad = int(request.form.get(f'cantidad_{prod.id_producto}', 1))
                prod.cantidad = cantidad
                total += (prod.precio or 0) * cantidad
                productos_a_pagar.append(prod)
    
    try:
        session['productos_boleta'] = json.loads(json.dumps(productos_a_pagar, default=decimal_default))
        session['total_boleta'] = total
        session['last_id_caja'] = id_caja
        
        # Debug
        if app.config.get('LOGIN_DEBUG'):
            print(f"[SESSION] total_boleta guardado: {total}")
            print(f"[SESSION] productos_boleta: {session['productos_boleta']}")
    except Exception as e:
        print(f"[ERROR] No se pudo guardar en session: {e}")
        flash('Hubo un error al procesar los productos.', 'danger')
        return redirect(request.referrer)

    # Guardamos productos y total en sesión (ya están guardados), y redirigimos
    # sólo con el id_caja para evitar exponer datos sensibles en la URL.
    return redirect(url_for('datos_cliente', id_caja=id_caja))


@app.route('/apertura', methods=['POST'])
def apertura_crear():
    if 'user_id' not in session:
        return redirect('/')
    id_caja = request.form.get('id_caja') or session.get('last_id_caja')
    try:
        saldo_inicio = int(request.form.get('saldo_inicio', 0))
    except Exception:
        saldo_inicio = 0

    # Verificar permisos del usuario para la caja solicitada
    try:
        permisos = Permiso.get_by_user_id(session['user_id'])
        allowed_caja_ids = [p.vta_cajas_id_caja for p in permisos]
    except Exception:
        allowed_caja_ids = []

    try:
        id_caja_int = int(id_caja)
    except Exception:
        flash('Caja inválida.', 'danger')
        return redirect(url_for('listar_aperturas'))

    if id_caja_int not in allowed_caja_ids:
        flash('No tienes permisos para abrir una apertura en esta caja.', 'danger')
        return redirect(url_for('listar_aperturas'))

    # Verificar que no exista apertura activa para la caja (regla: una apertura por caja)
    existing = Apertura.get_active_by_caja(id_caja_int)
    if existing:
        flash('Ya existe una apertura activa para esta caja.', 'warning')
        return redirect(url_for('listar_aperturas'))

    id_ap = Apertura.open_with_amount(id_caja_int, session['user_id'], saldo_inicio)
    if id_ap:
        # flash(f'Caja abierta (id {id_ap}).', 'success')
        pass
    else:
        flash('No se pudo abrir la caja.', 'danger')
    return redirect(url_for('listar_aperturas'))


@app.route('/apertura/<int:id_apertura>/cerrar', methods=['POST'])
def apertura_cerrar(id_apertura):
    if 'user_id' not in session:
        return redirect('/')
    # Si no se entrega `saldo_cierre`, usamos la suma de ventas como saldo final.
    observaciones = request.form.get('observaciones')

    try:
        # calcular totales de ventas para esta apertura
        total = Apertura.get_totals_for_apertura(id_apertura)

        saldo_cierre_raw = request.form.get('saldo_cierre')
        try:
            if saldo_cierre_raw is None or str(saldo_cierre_raw).strip() == '':
                saldo_cierre = total
            else:
                saldo_cierre = int(saldo_cierre_raw)
        except Exception:
            saldo_cierre = total

        diferencias = saldo_cierre - total

        # Intentar cerrar con resumen; capturamos excepciones para mostrar mensajes útiles
        try:
            res = Apertura.close_with_summary(id_apertura, saldo_cierre, total, diferencias, observaciones)
        except Exception as e:
            # Loggear en debug y notificar al usuario
            if app.config.get('LOGIN_DEBUG'):
                print(f"[APERTURA CLOSE ERROR] Error ejecutando close_with_summary for id={id_apertura}: {e}")
            flash('Error al cerrar la apertura (detalle en logs).', 'danger')
            return redirect(request.referrer or url_for('listar_aperturas'))

        if res is not False and res is not None:
            # flash('Caja cerrada correctamente.', 'success')
            # Guardar id para mostrar resumen automáticamente en la lista
            try:
                session['apertura_resumen_id'] = id_apertura
            except Exception:
                pass
            return redirect(url_for('listar_aperturas'))
        else:
            flash('No se pudo actualizar la apertura. Ver logs.', 'danger')
            return redirect(request.referrer or url_for('listar_aperturas'))

    except Exception as e:
        # Error inesperado durante el proceso
        if app.config.get('LOGIN_DEBUG'):
            print(f"[APERTURA CLOSE EXCEPTION] id={id_apertura} err={e}")
        flash(f'Error inesperado al intentar cerrar la apertura: {e}', 'danger')
        return redirect(request.referrer or url_for('listar_aperturas'))


@app.route('/apertura/<int:id_apertura>/resumen')
def resumen_apertura(id_apertura):
    if 'user_id' not in session:
        return redirect('/')
    ap = Apertura.get_by_id(id_apertura)
    if not ap:
        flash('Apertura no encontrada.', 'warning')
        return redirect(url_for('listar_aperturas'))

    total = Apertura.get_totals_for_apertura(id_apertura)

    # preparar contexto
    caja = Caja.get_by_id(ap.id_caja_fk) if hasattr(Caja, 'get_by_id') else None
    return render_template('apertura_resumen.html', apertura=ap, total=total, caja=caja)


@app.route('/apertura/<int:id_apertura>/resumen_fragment')
def resumen_apertura_fragment(id_apertura):
    if 'user_id' not in session:
        return ('', 401)
    ap = Apertura.get_by_id(id_apertura)
    if not ap:
        return ('', 404)
    total = Apertura.get_totals_for_apertura(id_apertura)
    caja = Caja.get_by_id(ap.id_caja_fk) if hasattr(Caja, 'get_by_id') else None
    return render_template('apertura_resumen_modal.html', apertura=ap, total=total, caja=caja)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/datos_cliente', methods=['GET', 'POST'])
def datos_cliente():
    if 'user_id' not in session:
        return redirect('/')

    if request.method == 'POST':
        # Validación: si el medio requiere voucher, asegúrese de que esté presente y sea numérico
        medio = request.form.get('medio_pago')
        if medio:
            medio = medio.strip()
        
        voucher = request.form.get('voucher')

        # Auto-generar voucher para transferencia
        if medio == 'transferencia':
            voucher = datetime.now().strftime('%Y%m%d')

        # Solo validamos voucher para debito y credito, ya que transferencia se autogenera
        medios_con_voucher = {'debito', 'credito'}
        if medio in medios_con_voucher:
            if not voucher or str(voucher).strip() == '':
                flash(f'Debe ingresar el número de voucher para el medio de pago seleccionado ({medio}).', 'danger')
                return render_template('procesamiento_pago.html', **request.form)
            if not str(voucher).strip().isdigit():
                flash('El número de voucher debe contener sólo dígitos.', 'danger')
                return render_template('procesamiento_pago.html', **request.form)
            # Limitar longitud para evitar valores fuera de rango en la BD
            if len(str(voucher).strip()) > 12:
                flash('El número de voucher es demasiado largo (máximo 12 dígitos).', 'danger')
                return render_template('procesamiento_pago.html', **request.form)

        # Guardar datos del cliente temporalmente en sesión. Todos los campos son opcionales.
        session['cliente_temp'] = {
            'nombre': request.form.get('nombre') or None,
            'rut': request.form.get('rut') or None,
            'correo': request.form.get('correo') or None,
            'telefono': request.form.get('telefono') or None,
            'medio_pago': medio,
            'voucher': voucher
        }

        # En lugar de redirigir a una confirmación intermedia, procesamos la venta
        # inmediatamente y mostramos el comprobante. Este flujo envía automáticamente
        # el comprobante por correo si existe `correo_cli`.
        # Reutilizamos la lógica previamente en /confirmar_pago POST.
        id_caja = request.form.get('id_caja') or session.get('last_id_caja')
        total = session.get('total_boleta', 0)
        correo = request.form.get('correo')
        nombre = request.form.get('nombre')
        productos_list = session.get('productos_boleta', [])

        if not productos_list:
            flash('Error: No se encontraron productos para procesar el pago.', 'warning')
            return redirect(url_for('ver_caja', id_caja=int(id_caja)))

        try:
            # 1. Verificar Apertura: buscamos una apertura activa para la caja.
            try:
                id_caja_int = int(id_caja)
            except Exception:
                flash('Caja inválida.', 'danger')
                return redirect(url_for('ver_caja', id_caja=int(id_caja)))

            # Verificar que el usuario tenga permiso para operar en esta caja
            permisos = Permiso.get_by_user_id(session['user_id'])
            allowed_caja_ids = [p.vta_cajas_id_caja for p in permisos]
            if id_caja_int not in allowed_caja_ids:
                flash('No tienes permiso para operar en esa caja.', 'danger')
                return redirect(url_for('ver_caja', id_caja=id_caja_int))

            # Buscar apertura activa por caja (regla: una apertura por caja)
            active_apertura = Apertura.get_active_by_caja(id_caja_int)
            if not active_apertura:
                flash('No hay una apertura activa para esta caja. Abre una en Gestión de Aperturas.', 'danger')
                return redirect(url_for('ver_caja', id_caja=id_caja_int))

            # 2. Registrar/obtener cliente en MySQL
            cliente_temp = session.get('cliente_temp', {})
            nombre_cli = cliente_temp.get('nombre') or nombre
            rut_cli = cliente_temp.get('rut')
            correo_cli = cliente_temp.get('correo') or correo
            telefono_cli = cliente_temp.get('telefono')

            # Buscar cliente por correo
            find_q = "SELECT id_cliente FROM vta_clientes WHERE email_cliente = %(email)s"
            found = connectToMySQL('sistemas').query_db(find_q, {'email': correo_cli})
            if found and isinstance(found, list) and len(found) > 0:
                id_cliente_fk = found[0].get('id_cliente')
                try:
                    # flash(f'Cliente existente usado (id {id_cliente_fk}).', 'info')
                    pass
                except Exception:
                    pass
            else:
                ins_q = "INSERT INTO vta_clientes (email_cliente, nombre_cliente, telefono_cliente) VALUES (%(email)s, %(nombre)s, %(telefono)s)"
                id_cliente_fk = connectToMySQL('sistemas').query_db(ins_q, {'email': correo_cli, 'nombre': nombre_cli or 'Cliente', 'telefono': telefono_cli})
                try:
                    if correo_cli:
                        # flash(f'Cliente creado (id {id_cliente_fk}) con correo {correo_cli}.', 'success')
                        pass
                    else:
                        # flash(f'Cliente creado (id {id_cliente_fk}) sin correo.', 'success')
                        pass
                except Exception:
                    pass

            # 3. Crear la Venta (con referencia al cliente)
            # Determinar lista de precio según la caja
            id_listaprecio = 168 if id_caja_int == 6 else 176
            
            # Agregar id_listaprecio a cada producto
            for prod in productos_list:
                prod['id_listaprecio'] = id_listaprecio
            
            data_venta = {
                'total_ventas': total,
                'id_apertura': active_apertura.id_apertura,
                'envio_correo': 1,  # Por defecto 1, se cambiará a 0 solo si hay error al enviar
                'id_cliente_fk': id_cliente_fk,
                'id_correlativo_flex': 0, # Valor por defecto para nuevo campo requerido
                'envio_boleta': 0         # Valor por defecto
            }

            id_venta = Venta.create(data_venta, productos_list)

            if not id_venta:
                flash('Hubo un error al registrar la venta en la base de datos.', 'danger')
                return redirect(url_for('ver_caja', id_caja=int(id_caja)))

            # 4. Registrar medio de pago en vta_mediopago
            medio = cliente_temp.get('medio_pago') or request.form.get('medio_pago')
            voucher_val = cliente_temp.get('voucher') or request.form.get('voucher')
            id_voucher_val = 0  # Inicializar por defecto para evitar UnboundLocalError
            
            try:
                if medio:
                    try:
                        id_voucher_val = int(voucher_val) if voucher_val is not None and str(voucher_val).strip() != '' else 0
                    except Exception:
                        id_voucher_val = 0

                    mp_q = "INSERT INTO vta_mediopago (tipo_pago, id_voucher, id_ventas_fk) VALUES (%(tipo)s, %(id_voucher)s, %(id_venta)s)"
                    connectToMySQL('sistemas').query_db(mp_q, {'tipo': medio, 'id_voucher': id_voucher_val, 'id_venta': id_venta})
            except Exception as e:
                if app.config.get('LOGIN_DEBUG'):
                    print(f"[PAYMENT DEBUG] Error insertando medio de pago: {e}")

            # flash(f'Venta #{id_venta} registrada con éxito!', 'success')

            # --- INTEGRACION FACTURA-X DESHABILITADA ---
            skip_api = True  # Deshabilitado para todas las cajas
            
            pdf_link = None
            api_success = False
            
            # 1. Preparar RUT
            rut_final = "66666666-6"
            if rut_cli:
                clean_rut = rut_cli.replace('.', '').replace(' ', '').upper()
                if '-' in clean_rut:
                    rut_final = clean_rut
                elif len(clean_rut) > 1:
                    rut_final = f"{clean_rut[:-1]}-{clean_rut[-1]}"
                if "NAN" in rut_final:
                    rut_final = "66666666-6"
            
            # 2. Preparar Items
            items_api = []
            for idx, prod in enumerate(productos_list, 1):
                cantidad = float(prod.get("cantidad", 1))
                precio = float(prod.get("precio") or prod.get("PRECIO") or 0)
                prod_name = prod.get("nombre") or prod.get("NOMBRE") or prod.get("name") or prod.get("NAME") or "Item Sin Nombre"
                amount = precio * cantidad
                
                items_api.append({
                    "line": str(idx),
                    "name": str(prod_name)[:40],
                    "quantity": str(int(cantidad)),
                    "price": str(int(precio)),
                    "amount": str(int(amount))
                })
            
            total_str = str(int(float(total))) if total else "0"
            
            # 3. Construir JSON (solo si no es caja variable)
            if not skip_api:
                now_dt = datetime.now()
                factura_json = {
                    "document_type": "CL39", 
                    "test": True, 
                    "numbering": False, 
                    "document": {
                        "number": "5001",
                        "currency": "CLP",
                        "issued_date": now_dt.strftime("%Y-%m-%d"),
                        "issued_date_time": now_dt.strftime("%Y-%m-%dT%H:%M:%S"),
                        "paymentTermsDescription": "CONTADO",
                        "service_type": "3",
                        "supplier": {
                            "tax_id": "99522880-7", 
                            "name": "ALIMENTAR SA",
                            "address": "AV QUILIN 1561",
                            "municipality": "QUILIN",
                            "city": "SANTIAGO"
                        },
                        "customer": {
                            "tax_id": rut_final,
                            "name": nombre_cli or "Cliente",
                            "activity": "COMERCIAL",
                            "contactEmail": correo_cli if correo_cli else "",
                            "contactPhone": "", 
                            "address": "",
                            "municipality": "",
                            "city": ""
                        },
                        "amount": {
                            "net": total_str,
                            "exempt": "0",
                            "vat_rate": "19.00",
                            "vat": "0",
                            "total": total_str
                        },
                        "taxes": [], 
                        "items": items_api
                    },
                    "custom": {
                        "Observaciones": "VENTA BOLETA",
                        "MntTotalWords": "" 
                    }
                }

                # 4. Llamar API
                API_URL = "https://services.factura-x.com/generation/cl/39" 
                API_TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJleHAiOjE3OTUyNjI0MDAsImlhdCI6MTc2MzczOTM4MiwianRpIjoiNlpVRWNwa3hkYVBTZDhjbkR1R3RLdyIsInVzZXJfaWQiOjMzLCJjb20uYXh0ZXJvaWQuaXNfc3RhZmYiOmZhbHNlLCJjb20uYXh0ZXJvaWQud29ya3NwYWNlcyI6eyJhY2NfTmNJTVBWa3FybTFnc1A2NkVRIjoiOTY4ODI3NTAtMiIsImFjY181b0p0aUt2NldrNkN0UUx4dzciOiI5OTUyMjg4MC03IiwiYWNjX0xjbnlKWWpNTFJ4S2FUYmVmdyI6Ijc4ODYxMjAwLTEifSwiY29tLmF4dGVyb2lkLmFsbG93ZWRfaW50ZXJ2YWwiOjEwMDB9.eoVYS5KqiQyGGtbsMc3kNMreuB0Cnoz8HImNmcxQUZY" 
                WORKSPACE_ID = "acc_5oJtiKv6Wk6CtQLxw7"

                headers = {
                    'Authorization': f"Bearer {API_TOKEN}", 
                    'x-ax-workspace': WORKSPACE_ID,
                    'Content-Type': 'application/json'
                }

                try:
                    if app.config.get('LOGIN_DEBUG'):
                        print("[FACTURA-X] Enviando request...")
                    
                    response = requests.post(API_URL, json=factura_json, headers=headers, timeout=15)
                    
                    if response.status_code == 200 or response.status_code == 201:
                        api_data = response.json()
                        doc_id = None
                        if 'id' in api_data:
                            doc_id = api_data['id']
                        elif 'document' in api_data and 'id' in api_data['document']:
                            doc_id = api_data['document']['id']
                        
                        if 'document' in api_data and 'pdf_plot' in api_data['document']:
                            pdf_link = api_data['document']['pdf_plot']
                        
                        if not pdf_link and doc_id:
                            pdf_link = f"https://services.factura-x.com/documents/{doc_id}?format=pdf"
                        
                        if pdf_link:
                            api_success = True
                            # flash('Boleta electrónica generada correctamente.', 'success')
                        else:
                            flash('Boleta generada pero no se obtuvo link PDF.', 'warning')
                    else:
                        error_msg = response.text
                        try:
                            error_json = response.json()
                            if 'message' in error_json:
                                error_msg = error_json['message']
                                if "No available numbering" in error_msg:
                                    error_msg = "Error Crítico: Se han agotado los folios."
                        except:
                            pass
                        flash(f'Error Factura-X: {error_msg}', 'warning')
                        if app.config.get('LOGIN_DEBUG'):
                            print(f"[FACTURA-X ERROR] {response.status_code} {response.text}")

                except Exception as e:
                    flash(f'Error conectando con Factura-X: {e}', 'warning')
                    if app.config.get('LOGIN_DEBUG'):
                        print(f"[FACTURA-X EXCEPTION] {e}")

            # Enviar Correo automáticamente si existe correo_cli
            # Solo envía comprobante de pago, el agente enviará la boleta después
            email_sent = False
            email_error = None
            if correo_cli:
                try:
                    # Preparar contenido común (Texto y HTML del comprobante)
                    subject = f'Comprobante de Venta Club Recrear OPEN DAY'
                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    try:
                        total_num = int(float(total))
                    except Exception:
                        try:
                            total_num = int(total)
                        except Exception:
                            total_num = 0

                    if nombre_cli and str(nombre_cli).strip():
                        saludo = f"Estimado/a {nombre_cli}"
                    else:
                        saludo = "Estimado/a Cliente"

                    body = (
                        f"{saludo},\n\nGracias por su compra.\n\n"
                        f"Resumen:\n"
                        f"- Nº Venta: {id_venta}\n"
                        f"- Total: ${total_num:,}\n"
                        f"- Fecha: {now}\n\n"
                        f"Su boleta electrónica será enviada en un correo separado.\n\n"
                        f"Saludos,\nClub Recrear"
                    )
                    
                    # Render HTML voucher for email
                    html_body = render_template('email_voucher.html', 
                        fecha=now,
                        total=total_num,
                        id_venta=id_venta,
                        cliente={'nombre': nombre_cli, 'correo': correo_cli},
                        medio_pago=medio,
                        id_voucher=id_voucher_val
                    )

                    # NO adjuntar PDF - el agente lo enviará después
                    attachments_list = []

                    # Enviar usando la función auxiliar send_email (solo comprobante)
                    send_email(correo_cli, subject, body, html_body=html_body, attachments=attachments_list)
                    
                    email_sent = True
                    # flash(f'Comprobante de compra enviado a {correo_cli}', 'success')

                except Exception as e:
                    email_error = str(e)
                    if app.config.get('LOGIN_DEBUG'):
                        print(f"[EMAIL ERROR] Error enviando email a {correo_cli}: {e}")
                    flash(f'No se pudo enviar el correo a {correo_cli}. Error: {e}', 'warning')

        except Exception as e:
            flash(f'Error inesperado durante el registro de la venta: {e}', 'danger')
            print(f"[SALE ERROR] {e}")
            return redirect(url_for('ver_caja', id_caja=int(id_caja)))

        # Preparar datos para mostrar comprobante en pantalla
        try:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            try:
                total_num = int(float(total))
            except Exception:
                try:
                    total_num = int(total)
                except Exception:
                    total_num = 0

            cliente_info = {
                'nombre': nombre_cli,
                'correo': correo_cli,
                'telefono': telefono_cli
            }

            comprobante_ctx = {
                'id_venta': id_venta,
                'medio_pago': medio,
                'id_voucher': id_voucher_val if 'id_voucher_val' in locals() else 0,
                'fecha': now,
                'total': total_num,
                'productos': productos_list,
                'cliente': cliente_info,
                'email_sent': email_sent,
                'email_error': email_error
            }
        except Exception as e:
            if app.config.get('LOGIN_DEBUG'):
                print(f"[COMPROBANTE DEBUG] Error preparando contexto del comprobante: {e}")
            comprobante_ctx = {
                'id_venta': id_venta,
                'medio_pago': medio,
                'id_voucher': 0,
                'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total': total,
                'productos': productos_list,
                'cliente': {'nombre': nombre_cli, 'correo': correo_cli},
                'email_sent': email_sent,
                'email_error': email_error
            }

        # Limpiar sesión después de generar comprobante
        session.pop('productos_boleta', None)
        session.pop('total_boleta', None)
        session.pop('cliente_temp', None)

        return render_template('comprobante.html', **comprobante_ctx)
    # Para GET: preferimos tomar la lista de productos y total desde la sesión
    # (más seguro) si no vienen en los query params.
    id_caja = request.args.get('id_caja')
    productos = request.args.get('productos')
    total = request.args.get('total')

    if not productos:
        try:
            productos = json.dumps(session.get('productos_boleta', []))
        except Exception:
            productos = '[]'

    if not total:
        total = session.get('total_boleta', 0)

    return render_template('procesamiento_pago.html', id_caja=id_caja, productos=productos, total=total)

# La ruta /confirmar_pago fue eliminada: el flujo ahora procesa la venta directamente
# desde /datos_cliente (POST) y renderiza `comprobante.html`.


@app.route('/apertura/<int:id_apertura>/export')
def export_apertura_xlsx(id_apertura):
    """Exportar las ventas de una apertura como archivo Excel (.xlsx) con formato.
    Columnas: FechaHora, NroVenta, MedioPago, Voucher, Total
    """
    if 'user_id' not in session:
        return redirect('/')

    # Select ventas and payment info. Avoid referencing potentially-missing columns
    # in SQL (previous COALESCE caused Unknown column errors). We'll determine
    # the date column in Python from the returned row keys.
    q = (
        "SELECT v.*, mp.tipo_pago AS medio_pago, mp.id_voucher AS voucher "
        "FROM vta_ventas v "
        "LEFT JOIN vta_mediopago mp ON mp.id_ventas_fk = v.id_ventas "
        "WHERE v.id_apertura = %(id_apertura)s ORDER BY v.id_ventas ASC;"
    )
    rows = connectToMySQL('sistemas').query_db(q, {'id_apertura': id_apertura})

    # Debug logging to diagnose empty exports. Prints SQL, apertura id and row count.
    try:
        try:
            cnt = len(rows) if rows is not None else 0
        except Exception:
            cnt = 'unknown'
        if app.config.get('LOGIN_DEBUG'):
            print(f"[EXPORT XLSX DEBUG] id_apertura={id_apertura} rows_returned={cnt}")
            try:
                print(f"[EXPORT XLSX DEBUG] sample_rows={rows[:3]}")
            except Exception:
                pass
        else:
            # Lightweight informational print to help the user when they run the server.
            print(f"[EXPORT XLSX] id_apertura={id_apertura} rows_returned={cnt}")
    except Exception:
        pass

    # Obtener la fecha de la apertura para usarla como fallback si la venta no tiene fecha
    try:
        apertura_obj = Apertura.get_by_id(id_apertura)
        apertura_date = getattr(apertura_obj, 'fecha_inicio_apertura', None) if apertura_obj is not None else None
    except Exception:
        apertura_date = None

    # Crear workbook en memoria usando openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        flash('Dependency missing: openpyxl no está instalado. Instale la dependencia y reinicie.', 'danger')
        return redirect(url_for('resumen_apertura', id_apertura=id_apertura))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Arqueo {id_apertura}"

    headers = ['FechaHora', 'NroVenta', 'MedioPago', 'Voucher', 'Total']
    header_font = Font(bold=True, color='FFFFFFFF')
    header_fill = PatternFill('solid', fgColor='2E8B57')  # verde
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='FFAAAAAA')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Escribir filas
    row_num = 2
    for r in rows or []:
        # Determine a sensible date field from available columns returned by DB.
        fecha_val = (
            r.get('fecha') or r.get('fecha_venta') or r.get('fecha_creacion') or
            r.get('created_at') or r.get('timestamp') or r.get('fecha_hora') or r.get('fechaVenta')
        )

        # Si no hay fecha en la venta, usar la fecha de la apertura (requerimiento solicitado)
        if not fecha_val:
            fecha_val = apertura_date
        # Parsear string a datetime si es necesario
        if isinstance(fecha_val, str) and fecha_val.strip():
            try:
                from datetime import datetime as _dt
                try:
                    fecha_dt = _dt.fromisoformat(fecha_val)
                except Exception:
                    fecha_dt = None
                if fecha_dt is None:
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d"):
                        try:
                            fecha_dt = _dt.strptime(fecha_val, fmt)
                            break
                        except Exception:
                            fecha_dt = None
                if fecha_dt:
                    fecha_val = fecha_dt
            except Exception:
                pass

        nro = r.get('id_ventas') or r.get('id_venta') or ''
        medio = r.get('medio_pago') or r.get('tipo_pago') or ''
        voucher = r.get('voucher') or ''
        if not voucher or str(voucher).strip() == '0':
            voucher = ''
        total = r.get('total_ventas') or r.get('total') or 0

        fcell = ws.cell(row=row_num, column=1, value=fecha_val)
        ws.cell(row=row_num, column=2, value=nro)
        ws.cell(row=row_num, column=3, value=medio)
        ws.cell(row=row_num, column=4, value=voucher)
        tcell = ws.cell(row=row_num, column=5, value=total)

        if fcell.value and not isinstance(fcell.value, str):
            fcell.number_format = 'DD-MM-YYYY HH:MM:SS'
        tcell.number_format = '#,##0'
        tcell.alignment = Alignment(horizontal='right')

        row_num += 1

    # Ajustar anchos y formato final
    col_widths = [20, 12, 18, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:E{row_num-1}"
    ws.freeze_panes = 'A2'

    import io as _io
    bio = _io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"arqueo_{id_apertura}.xlsx"
    resp = Response(bio.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp